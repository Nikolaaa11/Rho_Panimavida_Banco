# -*- coding: utf-8 -*-
"""Panimavida v8 - Ventana de Arbitraje: calculadora para el Comite de Credito."""
import sys, os
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference, Series

import os as _os
_BASE = _os.path.dirname(_os.path.abspath(__file__))
_DATOS = _os.path.join(_BASE, 'datos')


INK = 'FF1D1D1F'; GREY = 'FF6E6E73'; HAIR = 'FFE6E6EA'
GREEN = 'FF218358'; GREENL = 'FF30A46C'; RED = 'FFE5484D'; AMBER = 'FFB25E09'; BLUE = 'FF0B5FFF'


def F(s=10.5, b=False, c=INK, i=False):
    return Font(name='Segoe UI', size=s, bold=b, color=c, italic=i)


def fill(h):
    return PatternFill('solid', fgColor=h)


thin = Side(style='thin', color=HAIR)
BOT = Border(bottom=Side(style='thin', color=GREEN))
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()


def head(ws, title, sub, ncol=9):
    ws['A1'] = title
    ws['A1'].font = F(17, True)
    ws['A2'] = sub
    ws['A2'].font = F(9.5, c=GREY)
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 34
    ws['A2'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
    ws.sheet_view.showGridLines = False


def tblhead(ws, row, labels, widths=None):
    for j, l in enumerate(labels, 1):
        c = ws.cell(row=row, column=j, value=l)
        c.font = F(9.5, True)
        c.border = BOT
        c.alignment = Alignment(wrap_text=True, vertical='bottom',
                                horizontal='center' if j > 1 else 'left')
    ws.row_dimensions[row].height = 30
    if widths:
        for j, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = w


def note(ws, row, txt, ncol=9, h=32):
    c = ws.cell(row=row, column=1, value=txt)
    c.font = F(8.5, c=GREY, i=True)
    c.alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncol)
    ws.row_dimensions[row].height = h


def para(ws, row, txt, ncol, bold=False, color=INK, h=28, size=10):
    c = ws.cell(row=row, column=1, value=txt)
    c.font = F(size, bold, color)
    c.alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncol)
    ws.row_dimensions[row].height = h


YRS = list(range(2027, 2036))

# Direccion del break-even de referencia (340 ciclos, servicio US$400 mil) en la hoja Break_even.
# Es deterministica: la tabla arranca en la fila 6 y el servicio de 400.000 es la tercera fila.
# La columna E corresponde a 340 ciclos. Se verifica con un assert al final del build.
BE_ADDR = 'Break_even!$E$8'

# ============================================================ 1. PANEL
ws = wb.active
ws.title = 'Panel'
head(ws, 'Panimavida 13,2 kV - Ventana de arbitraje',
     'Calculadora de decision: operar a costo marginal (spot) vs. contratar. BESS 9 MW / 36 MWh. '
     'Ancla de mercado: ultimos 12 meses completos del nodo (23-may-2025 a 22-may-2026, 365 dias reales '
     'del Coordinador Electrico Nacional). Las celdas en azul son editables; el resto son formulas.')
for col, w in zip('ABCDEFGHI', [36, 15, 13, 15, 15, 15, 15, 15, 15]):
    ws.column_dimensions[col].width = w

r = 4
ws.cell(row=r, column=1, value='LA PREGUNTA DEL COMITE').font = F(11, True, GREEN)
r += 1
para(ws, r, 'Conviene ir a costo marginal en vez de firmar un PPA, y hasta cuando?', 9, h=18)
r += 2

ws.cell(row=r, column=1, value='RESPUESTA EN TRES CIFRAS').font = F(11, True, GREEN)
r += 1
cards = [
    ('Margen bruto - ultimos 12 meses reales', None, 'US$/ano',
     'Ciclando ~340 veces al ano sobre precios reales del nodo. No es proyeccion: es lo que el nodo pago.'),
    ('Spread necesario para pagar la deuda', None, 'USD/MWh',
     'A 340 ciclos y un servicio de deuda de US$400 mil. Contra un spread observado de 77,5 USD/MWh.'),
    ('Ano en que el escenario elegido cruza ese umbral', None, 'ano',
     'Antes de ese ano el spot domina. Ese es el borde de la ventana.'),
]
PANEL_CARD_ROWS = []
for lbl, fml, un, expl in cards:
    PANEL_CARD_ROWS.append(r)
    ws.cell(row=r, column=1, value=lbl).font = F(10, True)
    c = ws.cell(row=r, column=2, value=fml)
    c.font = F(15, True, GREEN)
    c.number_format = '#,##0'
    c.alignment = Alignment(horizontal='center')
    ws.cell(row=r, column=3, value=un).font = F(9, c=GREY)
    d = ws.cell(row=r, column=4, value=expl)
    d.font = F(8.5, c=GREY, i=True)
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=9)
    d.alignment = Alignment(wrap_text=True, vertical='center')
    ws.row_dimensions[r].height = 30
    r += 1
r += 1

ws.cell(row=r, column=1, value='PALANCAS EDITABLES').font = F(11, True, GREEN)
r += 1
tblhead(ws, r, ['Palanca', 'Valor', 'Unidad', 'Nota'])
r += 1
PS = r
levers = [
    ('Escenario activo (A / B / C)', 'B', '-',
     'A = ventana amplia. B = base. C = estres tipo NEM. Cambia toda la proyeccion.'),
    ('Ciclos completos por ano', 340, 'ciclos',
     '~0,93 por dia. En la ventana ancla, 88% de los dias tuvo spread mayor a 40 USD/MWh.'),
    ('CAPEX total', None, 'MUSD', 'REEMPLAZAR con el dato real del proyecto.'),
    ('Deuda / CAPEX', None, '%', 'REEMPLAZAR con el gearing del term sheet.'),
    ('Tasa de interes', None, '%', 'REEMPLAZAR.'),
    ('Plazo de la deuda', None, 'anos', 'REEMPLAZAR.'),
    ('DSCR minimo exigido', None, 'x', 'REEMPLAZAR con el covenant del term sheet.'),
]
for lbl, val, un, nt in levers:
    ws.cell(row=r, column=1, value=lbl).font = F(10)
    c = ws.cell(row=r, column=2)
    if val is None:
        c.value = '-'
        c.font = F(11, True, RED)
    else:
        c.value = val
        c.font = F(11, True, BLUE)
    c.fill = fill('FFF0F6FF')
    c.border = BOX
    c.alignment = Alignment(horizontal='center')
    ws.cell(row=r, column=3, value=un).font = F(9, c=GREY)
    d = ws.cell(row=r, column=4, value=nt)
    d.font = F(8.5, c=GREY)
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=9)
    r += 1

ESC = f'Panel!$B${PS}'
CIC = f'Panel!$B${PS+1}'
CAPEX = f'Panel!$B${PS+2}'
GEAR = f'Panel!$B${PS+3}'
TASA = f'Panel!$B${PS+4}'
PLAZO = f'Panel!$B${PS+5}'
r += 1
note(ws, r, 'Los cinco supuestos financieros estan vacios a proposito. Ninguna cifra de DSCR de este libro '
     'es valida hasta que se reemplacen por los terminos reales del term sheet. El resto del modelo '
     '-margen, spread, break-even y escenarios- no depende de ellos y ya es valido.')
r += 2

ws.cell(row=r, column=1, value='LO QUE ESTE LIBRO DEMUESTRA').font = F(11, True, GREEN)
r += 1
for t in [
    '1. El precio de venta nocturno NO viene cayendo: esta plano en 98-100 USD/MWh desde hace tres anos '
    '(ventanas de 12 meses completos: 108,2 / 98,2 / 99,6).',
    '2. Lo que comprimio el spread fue el COSTO DE CARGA, que subio de 10,1 a 22,1 USD/MWh en la ultima '
    'ventana, y que oscila con la hidrologia sin tendencia monotona (8,1 en 2024; 24,0 en 2025; 9,6 en ene-may 2026).',
    '3. Kimal-Lo Aguirre actua sobre el costo marginal DIURNO, es decir sobre el lado que hoy es el riesgo '
    'real, y lo empuja a la baja. En la evidencia disponible esa linea es viento de cola, no de frente.',
    '4. En 1.586 dias de registro, el precio de venta de las 4 horas de punta estuvo en o sobre 50 USD/MWh '
    'el 99,2% de los dias.',
]:
    para(ws, r, t, 9, h=32)
    r += 1

# ============================================================ 2. SUPUESTOS
ws = wb.create_sheet('Supuestos')
head(ws, 'Supuestos', 'Azul = editable. Cada supuesto de mercado esta anclado en dato real del nodo o en '
     'fuente citada. Rojo = falta el dato y bloquea el calculo de DSCR.', 4)
for col, w in zip('ABCD', [42, 16, 14, 66]):
    ws.column_dimensions[col].width = w
r = 4
tblhead(ws, r, ['Supuesto', 'Valor', 'Unidad', 'Fuente / justificacion'])
r += 1
S = {}
sup = [
    ('TECNICOS', None, None, None),
    ('Potencia del BESS', 9, 'MW', 'Ficha de proyecto'),
    ('Capacidad de almacenamiento', 36, 'MWh', 'Ficha de proyecto (duracion 4 h)'),
    ('Energia de carga por ciclo', 36, 'MWh', 'Igual a la capacidad'),
    ('Eficiencia round-trip', 0.83, '-', 'Conservadora frente al 88% nominal'),
    ('Energia de venta por ciclo', 'FML_VENTA', 'MWh', 'Carga x eficiencia'),
    ('Degradacion anual de capacidad', 0.02, '/ano', 'Supuesto LFP conservador; ajustar con la garantia del proveedor'),
    ('ANCLA DE MERCADO - ultimos 12 meses completos', None, None, None),
    ('Costo de carga (nivel)', 22.1, 'USD/MWh',
     '365 dias reales, 23-may-2025 a 22-may-2026. Incluye la hidrologia alta de 2025, por lo que es conservador'),
    ('Precio de venta 4 h punta (nivel)', 99.6, 'USD/MWh', 'Misma ventana de 365 dias'),
    ('Spread', 'FML_SPREAD', 'USD/MWh', 'Venta menos carga'),
    ('Piso de venta observado', 50, 'USD/MWh',
     'El 99,2% de los 1.586 dias del registro estuvo en o sobre este nivel'),
    ('Techo de carga en el modelo', 40, 'USD/MWh',
     'Limite superior asumido. El maximo anual observado fue 45,0 USD/MWh en 2022'),
    ('OPERACION', None, None, None),
    ('Umbral de spread para ciclar', 20, 'USD/MWh',
     'Bajo este umbral no se cicla. El margen es casi insensible al umbral entre 0 y 30'),
    ('OPEX anual', None, 'kUSD/ano',
     'O&M, servicio de pronostico day-ahead, seguros y cargos del Coordinador. REEMPLAZAR'),
    ('Reserva de degradacion', None, 'kUSD/ano', 'Augmentation o reposicion de celdas. REEMPLAZAR'),
]
row_carga_ciclo = None
row_efic = None
for lbl, val, un, src in sup:
    if val is None and un is None:
        ws.cell(row=r, column=1, value=lbl).font = F(9.5, True, GREEN)
        r += 1
        continue
    ws.cell(row=r, column=1, value=lbl).font = F(10)
    c = ws.cell(row=r, column=2)
    if val is None:
        c.value = '-'
        c.font = F(10.5, True, RED)
    elif val == 'FML_VENTA':
        c.value = f'=$B${row_carga_ciclo}*$B${row_efic}'
        c.font = F(10.5, True)
        c.number_format = '0.00'
    elif val == 'FML_SPREAD':
        c.value = f'=$B${r-1}-$B${r-2}'
        c.font = F(10.5, True)
        c.number_format = '0.0'
    else:
        c.value = val
        c.font = F(10.5, True, BLUE)
        c.fill = fill('FFF0F6FF')
        if isinstance(val, float) and val < 1:
            c.number_format = '0%'
    c.border = BOX
    c.alignment = Alignment(horizontal='center')
    ws.cell(row=r, column=3, value=un).font = F(9, c=GREY)
    d = ws.cell(row=r, column=4, value=src)
    d.font = F(8.5, c=GREY)
    d.alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[r].height = 26
    if lbl == 'Energia de carga por ciclo':
        row_carga_ciclo = r
    if lbl == 'Eficiencia round-trip':
        row_efic = r
    S[lbl] = f'Supuestos!$B${r}'
    r += 1

MWH_C = S['Energia de carga por ciclo']
MWH_V = S['Energia de venta por ciclo']
V0 = S['Precio de venta 4 h punta (nivel)']
C0 = S['Costo de carga (nivel)']
PISO = S['Piso de venta observado']
TECHO = S['Techo de carga en el modelo']
DEGR = S['Degradacion anual de capacidad']

# ============================================================ 3. HISTORICO
ws = wb.create_sheet('Historico_Ancla')
head(ws, 'Historico del nodo - por que el ancla son 12 meses completos',
     'Un ano calendario parcial mezcla estacionalidad y no es comparable. Las ventanas de 12 meses completos '
     'si lo son, y son la base de la proyeccion. 1.586 dias reales del Coordinador Electrico Nacional.', 7)
for col, w in zip('ABCDEFG', [30, 13, 13, 13, 13, 13, 34]):
    ws.column_dimensions[col].width = w
r = 4
ws.cell(row=r, column=1, value='A. VENTANAS DE 12 MESES COMPLETOS - la serie limpia').font = F(11, True, GREEN)
r += 1
tblhead(ws, r, ['Ventana', 'Carga', 'Venta', 'Spread', 'Delta carga', 'Delta venta', 'Delta spread'])
r += 1
w12 = [('23-may-2022 a 22-may-2023', 37.7, 200.0, 162.3, None, None, None),
       ('23-may-2023 a 22-may-2024', 12.3, 108.2, 96.0, -25.4, -91.8, -66.4),
       ('23-may-2024 a 22-may-2025', 10.1, 98.2, 88.1, -2.1, -10.0, -7.9),
       ('23-may-2025 a 22-may-2026', 22.1, 99.6, 77.5, 12.0, 1.4, -10.6)]
for i, (lb, c1, v, sp, dc, dv, ds) in enumerate(w12):
    last = i == len(w12) - 1
    ws.cell(row=r, column=1, value=lb).font = F(10, last)
    for j, val in enumerate([c1, v, sp], 2):
        cc = ws.cell(row=r, column=j, value=val)
        cc.number_format = '0.0'
        cc.font = F(10.5, last)
        cc.alignment = Alignment(horizontal='center')
    for j, val in enumerate([dc, dv, ds], 5):
        if val is not None:
            cc = ws.cell(row=r, column=j, value=val)
            cc.number_format = '+0.0;-0.0'
            bad = (j == 5 and val > 0) or (j > 5 and val < 0)
            cc.font = F(9.5, False, RED if bad else GREENL)
            cc.alignment = Alignment(horizontal='center')
    if last:
        ws.cell(row=r, column=7, value='ANCLA DEL MODELO').font = F(9, True, GREEN)
    r += 1
r += 1
para(ws, r, 'LECTURA CENTRAL - el lado venta esta plano; el que se movio es el lado carga:', 7, True, GREEN, 18)
r += 1
for t in ['Venta: 108,2 / 98,2 / 99,6 USD/MWh en las ultimas tres ventanas. Variacion neta de -8,6 en dos anos, y +1,4 en la ultima.',
          'Carga: 12,3 / 10,1 / 22,1 USD/MWh. Subio 12,0 en la ultima ventana, y eso explica practicamente toda la compresion reciente del spread.',
          'Implicancia: el riesgo del proyecto no es que caiga el precio de venta nocturno. Es que suba el costo de cargar de dia.']:
    para(ws, r, t, 7, h=26)
    r += 1
r += 1
ws.cell(row=r, column=1, value='B. LA CARGA TIENE TENDENCIA O ES HIDROLOGIA?').font = F(11, True, GREEN)
r += 1
tblhead(ws, r, ['Ano calendario', 'Carga', 'Venta', 'Spread', 'Carga ene-may', '', 'Nota'])
r += 1
for a, c1, v, sp, cm, nt in [(2022, 45.0, 184.8, 139.8, 49.1, 'Previo a la saturacion solar'),
                             (2023, 19.9, 143.1, 123.3, 31.4, 'Pico de la crisis energetica'),
                             (2024, 8.1, 93.8, 85.7, 13.4, 'Carga minima del registro'),
                             (2025, 24.0, 113.0, 89.0, 15.6, 'Hidrologia alta: la carga sube'),
                             (2026, 9.6, 85.0, 75.4, 9.6, 'Parcial (ene a 22-may). La carga vuelve a bajar')]:
    ws.cell(row=r, column=1, value=a).font = F(10)
    for j, val in enumerate([c1, v, sp, cm], 2):
        cc = ws.cell(row=r, column=j, value=val)
        cc.number_format = '0.0'
        cc.font = F(10.5)
        cc.alignment = Alignment(horizontal='center')
    d = ws.cell(row=r, column=7, value=nt)
    d.font = F(8.5, c=GREY)
    d.alignment = Alignment(wrap_text=True, vertical='top')
    r += 1
r += 1
note(ws, r, 'La carga oscila entre 8,1 y 45,0 USD/MWh sin tendencia monotona: 8,1 (2024), 24,0 (2025), '
     '9,6 (ene-may 2026). El alza de 2025 es consistente con mayor hidrologia -menos vertimiento, menos horas '
     'a costo cero- y ya revirtio. El modelo trata el alza del costo de carga como riesgo estructural de largo '
     'plazo, no como tendencia ya instalada.', 7)
r += 2
ws.cell(row=r, column=1, value='C. PISO DEL PRECIO DE VENTA - registro completo, 1.586 dias').font = F(11, True, GREEN)
r += 1
tblhead(ws, r, ['Umbral', 'Dias bajo el umbral', '% del registro', '% en o sobre el umbral', '', '', ''])
r += 1
for u, n, p in [(40, 4, 0.3), (50, 12, 0.8), (55, 30, 1.9), (60, 87, 5.5), (70, 263, 16.6)]:
    k = u == 50
    ws.cell(row=r, column=1, value=f'venta menor a {u} USD/MWh').font = F(10, k)
    cc = ws.cell(row=r, column=2, value=n)
    cc.font = F(10.5, k)
    cc.alignment = Alignment(horizontal='center')
    cc = ws.cell(row=r, column=3, value=p / 100)
    cc.number_format = '0.0%'
    cc.font = F(10.5, k)
    cc.alignment = Alignment(horizontal='center')
    cc = ws.cell(row=r, column=4, value=1 - p / 100)
    cc.number_format = '0.0%'
    cc.font = F(10.5, k, GREEN if k else INK)
    cc.alignment = Alignment(horizontal='center')
    r += 1
r += 1
note(ws, r, 'Precision obligatoria: el minimo absoluto del registro es 0,0 USD/MWh (30-oct-2022), un dia '
     'atipico. La afirmacion correcta y verificable es que el 99,2% de los 1.586 dias tuvo precio de venta en '
     'o sobre 50 USD/MWh. En la ventana ancla el peor dia fue 48,8 y el percentil 10 fue 58,5 USD/MWh.', 7)

# ============================================================ 4. ESCENARIOS
ws = wb.create_sheet('Escenarios')
head(ws, 'Escenarios de compresion - una sola variable',
     'La variable que manda es la velocidad con que el almacenamiento del SEN llega a comprimir el nodo del '
     'Maule. Se modela por los DOS lados, como advirtio el asesor tecnico: la venta baja y la carga sube.', 11)
for col, w in zip('ABCDEFGHIJK', [30, 13, 13, 13, 13, 13, 13, 13, 13, 13, 13]):
    ws.column_dimensions[col].width = w
r = 4
ws.cell(row=r, column=1, value='PARAMETROS DE CADA ESCENARIO (editables)').font = F(11, True, GREEN)
r += 1
tblhead(ws, r, ['Escenario', 'Ano inicio compresion', 'Delta venta anual', 'Delta carga anual', 'Justificacion'])
r += 1
P0 = r
escs = [('A. Ventana amplia', 2031, -2.0, 1.0,
         'El almacenamiento sigue concentrado en el norte (86% de los MW en construccion) y la CNE proyecta desacople del sur hacia 2032.'),
        ('B. Base', 2029, -4.0, 2.5,
         'Compresion gradual desde la entrada de Kimal-Lo Aguirre (2029-2030) y llegada parcial de BESS al centro-sur.'),
        ('C. Estres tipo NEM', 2027, -10.0, 5.0,
         'Replica el precedente australiano: spread -85% en un ano al pasar la flota BESS de 4.360 a 9.000 MW.')]
for lb, ai, dv, dc, j in escs:
    ws.cell(row=r, column=1, value=lb).font = F(10, True)
    for col, val, fmt in [(2, ai, '0'), (3, dv, '+0.0;-0.0'), (4, dc, '+0.0;-0.0')]:
        c = ws.cell(row=r, column=col, value=val)
        c.font = F(10.5, True, BLUE)
        c.fill = fill('FFF0F6FF')
        c.border = BOX
        c.number_format = fmt
        c.alignment = Alignment(horizontal='center')
    d = ws.cell(row=r, column=5, value=j)
    d.font = F(8.5, c=GREY)
    d.alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=11)
    ws.row_dimensions[r].height = 30
    r += 1
rows_esc = {'A': P0, 'B': P0 + 1, 'C': P0 + 2}
r += 1

blocks = [('PROYECCION - precio de venta (USD/MWh)', 'venta'),
          ('PROYECCION - costo de carga (USD/MWh)', 'carga'),
          ('PROYECCION - spread (USD/MWh)', 'spread')]
VROW, CROW, SROW = {}, {}, {}
catrow = None
for title, kind in blocks:
    ws.cell(row=r, column=1, value=title).font = F(11, True, GREEN)
    r += 1
    tblhead(ws, r, ['Escenario'] + [str(y) for y in YRS])
    if catrow is None:
        catrow = r
    r += 1
    for tag in ['A', 'B', 'C']:
        pr = rows_esc[tag]
        ws.cell(row=r, column=1, value=f'Escenario {tag}').font = F(10.5, kind == 'spread')
        for j, y in enumerate(YRS, 2):
            L = get_column_letter(j)
            if kind == 'venta':
                f = (f'=MAX({PISO},{V0}+IF({y}<$B${pr},0,({y}-$B${pr}+1)*$C${pr}))')
            elif kind == 'carga':
                f = (f'=MIN({TECHO},{C0}+IF({y}<$B${pr},0,({y}-$B${pr}+1)*$D${pr}))')
            else:
                f = f'={L}{VROW[tag]}-{L}{CROW[tag]}'
            c = ws.cell(row=r, column=j, value=f)
            c.number_format = '0.0'
            c.font = F(10.5, kind == 'spread')
            c.alignment = Alignment(horizontal='center')
        if kind == 'venta':
            VROW[tag] = r
        elif kind == 'carga':
            CROW[tag] = r
        else:
            SROW[tag] = r
        r += 1
    r += 1

note(ws, r, 'El piso de venta (50 USD/MWh) y el techo de carga (40 USD/MWh) acotan la proyeccion: el spread no '
     'cae indefinidamente porque el precio nocturno lo fija el costo variable de la ultima unidad termica que '
     'entra al anochecer, no el sol. Ese piso esta calibrado sobre el 99,2% del registro historico.', 11)
r += 2

ch = LineChart()
ch.title = 'Spread proyectado por escenario (USD/MWh)'
ch.height = 8.5
ch.width = 22
ch.y_axis.title = 'USD/MWh'
for tag in ['A', 'B', 'C']:
    ref = Reference(ws, min_col=2, max_col=1 + len(YRS), min_row=SROW[tag])
    ch.append(Series(ref, title=f'Escenario {tag}'))
ch.set_categories(Reference(ws, min_col=2, max_col=1 + len(YRS), min_row=catrow))
ws.add_chart(ch, f'A{r}')

# ============================================================ 5. MARGEN + DSCR
ws = wb.create_sheet('Margen_DSCR')
head(ws, 'Margen y cobertura de deuda',
     'El margen no depende de los supuestos financieros: se calcula sobre precios y volumenes, y ya es valido. '
     'El DSCR si depende, y esta bloqueado hasta que se ingresen los terminos reales del term sheet.', 11)
for col, w in zip('ABCDEFGHIJK', [36, 14, 14, 13, 13, 13, 13, 13, 13, 13, 13]):
    ws.column_dimensions[col].width = w
r = 4
ws.cell(row=r, column=1, value='A. VALIDACION CONTRA DATO REAL (sin proyeccion)').font = F(11, True, GREEN)
r += 1
tblhead(ws, r, ['Concepto', 'Valor', 'Unidad', 'Como se obtiene'])
r += 1
real = [('Ciclos en la ventana ancla', 340, 'ciclos',
         'Dias con spread mayor a 20 USD/MWh en los ultimos 12 meses completos'),
        ('Margen bruto de arbitraje', 814738, 'US$/ano',
         'Suma dia a dia de venta x 29,88 menos carga x 36, sobre los dias efectivamente ciclados'),
        ('Margen si se cicla todos los dias', 810174, 'US$/ano',
         '359 ciclos. Practicamente identico: el margen es robusto al criterio de ciclado'),
        ('Margen a 182 ciclos (plan conservador)', 524363, 'US$/ano',
         'Mitad de utilizacion, calculado sobre el ano mas comprimido del registro')]
MB_ROW = None
for i, (lb, v, un, how) in enumerate(real):
    k = i == 1
    ws.cell(row=r, column=1, value=lb).font = F(10, k)
    c = ws.cell(row=r, column=2, value=v)
    c.number_format = '#,##0'
    c.font = F(12 if k else 10.5, k, GREEN if k else INK)
    c.alignment = Alignment(horizontal='center')
    ws.cell(row=r, column=3, value=un).font = F(9, c=GREY)
    d = ws.cell(row=r, column=4, value=how)
    d.font = F(8.5, c=GREY)
    d.alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=11)
    ws.row_dimensions[r].height = 26
    if k:
        MB_ROW = r
    r += 1
r += 1
note(ws, r, 'UNIFICACION DE CIFRAS - el expediente actual tiene dos numeros distintos y conviene explicar la '
     'diferencia antes de que la encuentren: el informe v7 reporta US$493 mil/ano (precio-nivel anual por '
     'volumen, a 182 ciclos) y el correo enviado menciona cerca de US$0,9 millones/ano (seleccion dia a dia con '
     'ciclado diario). Ambos son correctos bajo su propio metodo. La cifra que este libro propone como titular '
     'es US$815 mil/ano: dia a dia, 340 ciclos, sobre los ultimos 12 meses completos y reales.', 11, 44)
r += 2

ws.cell(row=r, column=1, value='B. MARGEN BRUTO PROYECTADO POR ESCENARIO (US$/ano)').font = F(11, True, GREEN)
r += 1
tblhead(ws, r, ['Escenario'] + [str(y) for y in YRS])
r += 1
MROW = {}
for tag in ['A', 'B', 'C']:
    ws.cell(row=r, column=1, value=f'Escenario {tag}').font = F(10.5, True)
    for j, y in enumerate(YRS, 2):
        L = get_column_letter(j)
        n = y - 2027
        f = (f'={CIC}*(Escenarios!{L}{VROW[tag]}*{MWH_V}*(1-{DEGR})^{n}'
             f'-Escenarios!{L}{CROW[tag]}*{MWH_C})')
        c = ws.cell(row=r, column=j, value=f)
        c.number_format = '#,##0'
        c.font = F(10)
        c.alignment = Alignment(horizontal='center')
    MROW[tag] = r
    r += 1
r += 1
note(ws, r, 'Incluye degradacion de capacidad compuesta sobre el volumen de venta. No incluye OPEX ni reserva '
     'de degradacion: son margenes brutos de arbitraje, directamente comparables con la cifra real de la '
     'seccion A.', 11)
r += 2

ws.cell(row=r, column=1, value='C. COBERTURA DE DEUDA').font = F(11, True, GREEN)
r += 1
tblhead(ws, r, ['Concepto', 'Valor', 'Unidad', 'Nota'])
r += 1
ws.cell(row=r, column=1, value='Monto de deuda').font = F(10)
c = ws.cell(row=r, column=2, value=f'=IF(OR({CAPEX}="-",{GEAR}="-"),"faltan datos",{CAPEX}*{GEAR})')
c.font = F(10.5)
c.alignment = Alignment(horizontal='center')
ws.cell(row=r, column=3, value='MUSD').font = F(9, c=GREY)
ws.cell(row=r, column=4, value='CAPEX por gearing, tomados del Panel').font = F(8.5, c=GREY)
DEUDA = f'$B${r}'
r += 1
ws.cell(row=r, column=1, value='Servicio anual de la deuda').font = F(10)
c = ws.cell(row=r, column=2,
            value=f'=IF(OR({TASA}="-",{PLAZO}="-",{DEUDA}="faltan datos"),"faltan datos",'
                  f'-PMT({TASA},{PLAZO},{DEUDA})*1000000)')
c.number_format = '#,##0'
c.font = F(10.5)
c.alignment = Alignment(horizontal='center')
ws.cell(row=r, column=3, value='US$/ano').font = F(9, c=GREY)
ws.cell(row=r, column=4, value='Cuota constante. Ajustar si hay gracia o perfil sculpted').font = F(8.5, c=GREY)
SERV = f'$B${r}'
r += 1
ws.cell(row=r, column=1, value='DSCR con el margen real de la ventana ancla').font = F(10, True)
c = ws.cell(row=r, column=2, value=f'=IF({SERV}="faltan datos","faltan datos",$B${MB_ROW}/{SERV})')
c.number_format = '0.00"x"'
c.font = F(12, True, GREEN)
c.alignment = Alignment(horizontal='center')
ws.cell(row=r, column=4, value='Margen bruto sobre servicio. Restar OPEX y reserva para el DSCR definitivo').font = F(8.5, c=GREY)
r += 2

ws.cell(row=r, column=1, value='D. ANO EN QUE EL ESCENARIO CRUZA EL BREAK-EVEN').font = F(11, True, GREEN)
r += 1
ws.cell(row=r, column=1, value='Escenario activo (tomado del Panel)').font = F(10)
c = ws.cell(row=r, column=2, value=f'={ESC}')
c.font = F(10.5, True, BLUE)
c.alignment = Alignment(horizontal='center')
r += 1
sA, sB, sC = SROW['A'], SROW['B'], SROW['C']

# Filas auxiliares no-matriciales (evitan Ctrl+Shift+Enter y son legibles para el banco)
ws.cell(row=r, column=1, value='Ano').font = F(9.5, c=GREY)
AUX_Y = r
for j, y in enumerate(YRS, 2):
    c = ws.cell(row=r, column=j, value=y)
    c.font = F(9.5, c=GREY)
    c.alignment = Alignment(horizontal='center')
r += 1
ws.cell(row=r, column=1, value='Spread del escenario activo (USD/MWh)').font = F(10)
AUX_S = r
for j, y in enumerate(YRS, 2):
    L = get_column_letter(j)
    f = (f'=IF({ESC}="A",Escenarios!{L}{sA},IF({ESC}="B",Escenarios!{L}{sB},Escenarios!{L}{sC}))')
    c = ws.cell(row=r, column=j, value=f)
    c.number_format = '0.0'
    c.font = F(10)
    c.alignment = Alignment(horizontal='center')
r += 1
ws.cell(row=r, column=1, value='Bajo el break-even? (1 = si)').font = F(9.5, c=GREY)
AUX_F = r
for j, y in enumerate(YRS, 2):
    L = get_column_letter(j)
    c = ws.cell(row=r, column=j, value=f'=IF({L}{AUX_S}<{BE_ADDR},1,0)')
    c.font = F(9.5, c=GREY)
    c.alignment = Alignment(horizontal='center')
r += 2

lastcol = get_column_letter(1 + len(YRS))
ws.cell(row=r, column=1, value='Primer ano con spread bajo el break-even').font = F(10, True)
f = (f'=IFERROR(INDEX($B${AUX_Y}:${lastcol}${AUX_Y},MATCH(1,$B${AUX_F}:${lastcol}${AUX_F},0)),'
     f'"no cruza hasta {YRS[-1]}")')
c = ws.cell(row=r, column=2, value=f)
c.font = F(13, True, AMBER)
c.alignment = Alignment(horizontal='center')
ws.cell(row=r, column=3, value='borde de la ventana').font = F(9, c=GREY)
CRUCE = r
r += 2
note(ws, r, 'Este es el numero que responde la pregunta del Comite: hasta que ano el spot domina bajo el '
     'escenario elegido. Cambiando el escenario en el Panel se ve moverse el borde de la ventana. Las tres '
     'filas auxiliares de arriba dejan el calculo a la vista, sin formulas matriciales.', 11)

SRC_MARGEN = f'Margen_DSCR!$B${MB_ROW}'
SRC_CRUCE = f'Margen_DSCR!$B${CRUCE}'

# ============================================================ 6. BREAK-EVEN
ws = wb.create_sheet('Break_even')
head(ws, 'Break-even - cuanto tiene que caer el mercado antes de que duela',
     'Spread minimo necesario para cubrir el servicio de la deuda, segun numero de ciclos. Se lee al revés que '
     'un caso base: no dice cuanto se gana, dice cuanta holgura hay.', 7)
for col, w in zip('ABCDEFG', [30, 15, 15, 15, 15, 15, 26]):
    ws.column_dimensions[col].width = w
r = 4
ws.cell(row=r, column=1, value='Spread requerido (USD/MWh) por nivel de servicio de deuda').font = F(11, True, GREEN)
r += 1
tblhead(ws, r, ['Servicio de deuda (US$/ano)', '182 ciclos', '250 ciclos', '300 ciclos', '340 ciclos', '365 ciclos', ''])
r += 1
BE0 = r
for ds in [200000, 300000, 400000, 500000, 600000]:
    k = ds == 400000
    c = ws.cell(row=r, column=1, value=ds)
    c.number_format = '#,##0'
    c.font = F(10.5, k)
    for j, nc in enumerate([182, 250, 300, 340, 365], 2):
        f = f'=({ds}/{nc}+{C0}*{MWH_C})/{MWH_V}-{C0}'
        cc = ws.cell(row=r, column=j, value=f)
        cc.number_format = '0.0'
        cc.font = F(10.5, k)
        cc.alignment = Alignment(horizontal='center')
    if k:
        ws.cell(row=r, column=7, value='referencia del Panel').font = F(8.5, c=GREY)
        SRC_BE = f'Break_even!$E${r}'   # columna E = 340 ciclos
    r += 1
r += 1
para(ws, r, 'COMO SE LEE:', 7, True, GREEN, 18)
r += 1
for t in ['El spread observado en la ventana ancla (ultimos 12 meses completos) es 77,5 USD/MWh.',
          'A 340 ciclos, cubrir un servicio de deuda de US$400 mil requiere un spread de 43,9 USD/MWh.',
          'Es decir: el spread tendria que caer un 43% desde el nivel observado antes de comprometer el servicio de la deuda.',
          'En la ventana ancla el percentil 10 del spread diario fue 29,2 USD/MWh y el 88% de los dias supero 40 USD/MWh.']:
    para(ws, r, t, 7, h=22)
    r += 1
r += 1
note(ws, r, 'El calculo mantiene el costo de carga en el nivel de la ventana ancla (22,1 USD/MWh), que ya '
     'incorpora la hidrologia alta de 2025. Con el costo de carga de 2026 (9,6 USD/MWh) el spread requerido '
     'seria todavia menor.', 7)

# ============================================================ 7. GATILLOS
ws = wb.create_sheet('Gatillos')
head(ws, 'Sistema de gatillos - de apuesta a politica monitoreada',
     'Convierte la decision en una politica revisable trimestralmente. Cada gatillo tiene metrica, fuente, '
     'umbral y accion predefinida. Es la respuesta a la pregunta "y si se equivocan?".', 6)
for col, w in zip('ABCDEF', [5, 30, 30, 17, 17, 42]):
    ws.column_dimensions[col].width = w
r = 4
tblhead(ws, r, ['#', 'Gatillo', 'Metrica y fuente', 'Umbral', 'Valor actual', 'Accion al activarse'])
r += 1
gats = [('1', 'Compresion del spread', 'Spread movil 12M del nodo. Coordinador', 'menor a 55 USD/MWh', '77,5',
         'Acelerar la negociacion del contrato de venta o del floor nocturno'),
        ('2', 'Erosion del piso nocturno', 'Percentil 10 mensual del precio de punta 4 h. Coordinador', 'menor a 50 USD/MWh', '58,5',
         'Revisar el caso base y el perfil de amortizacion con el banco'),
        ('3', 'Alza del costo de carga', 'Costo de carga movil 12M. Coordinador', 'mayor a 35 USD/MWh', '22,1',
         'Reponderar el horizonte de la Fase 1. Es el riesgo que senalo el asesor tecnico'),
        ('4', 'BESS materializado en el centro-sur', 'MW en OPERACION en Maule, Nuble y Biobio. Coordinador', 'mayor a 300 MW', '0 MW registrado',
         'Reevaluar la exposicion locacional y adelantar la contratacion'),
        ('5', 'Refuerzo de transmision', 'Entrada en operacion de obra centro-sur o del HVDC. PET de la CNE', 'obra en servicio', 'Kimal-Lo Aguirre 2029-30',
         'Reevaluar el acoplamiento del nodo al troncal'),
        ('6', 'Cobertura de deuda', 'DSCR proyectado a 12 meses', 'bajo el covenant', 'faltan datos',
         'Activar contratacion parcial de la energia de venta')]
for n, g, m, u, v, a in gats:
    ws.cell(row=r, column=1, value=n).font = F(10, True, GREEN)
    ws.cell(row=r, column=2, value=g).font = F(10, True)
    for col, val in [(3, m), (4, u), (5, v), (6, a)]:
        c = ws.cell(row=r, column=col, value=val)
        c.font = F(9 if col in (3, 6) else 10, col == 5,
                   GREENL if col == 5 else (GREY if col in (3, 6) else INK))
        c.alignment = Alignment(wrap_text=True, vertical='top',
                                horizontal='center' if col in (4, 5) else 'left')
    ws.row_dimensions[r].height = 36
    r += 1
r += 1
note(ws, r, 'Los umbrales son propuestas calibradas sobre el registro historico y deben ratificarse con el '
     'banco; el del gatillo 6 requiere el covenant del term sheet. Frecuencia de revision: TRIMESTRAL, no '
     'anual. El precedente australiano -spread -85% en un ano- demuestra que la compresion puede ser abrupta.', 6)

# ============================================================ 8. INSTRUMENTOS
ws = wb.create_sheet('Instrumentos')
head(ws, 'Que contrato, y cuando - la distincion que falta en el expediente',
     'El expediente dice "no PPA ahora, reevaluar hacia 2030" y eso suena contradictorio. No lo es: son dos '
     'contratos distintos que hoy se llaman igual.', 3)
for col, w in zip('ABC', [32, 46, 46]):
    ws.column_dimensions[col].width = w
r = 4
tblhead(ws, r, ['', 'PPA de COMPRA a 28 USD/MWh (la oferta actual)', 'Instrumento de VENTA (Fase 2, 2030-2032)'])
r += 1
rows = [('Que fija', 'El precio al que el proyecto COMPRA su energia de carga', 'El precio al que el proyecto VENDE en la noche'),
        ('Riesgo que cubre', 'Que SUBA el precio de compra', 'Que CAIGA el precio de venta'),
        ('Es el riesgo real del proyecto?', 'No. El proyecto es posicion vendedora', 'Si. Es exactamente el riesgo del activo'),
        ('Efecto sobre el margen', 'Lo destruye: cerca de US$83 mil/ano de sobrecosto', 'Lo estabiliza'),
        ('Efecto sobre el DSCR', 'Lo deteriora: agrega costo fijo sin cubrir riesgo', 'Lo mejora: reduce la volatilidad de caja'),
        ('VEREDICTO', 'Rechazar de forma permanente, no "por ahora"', 'Negociar desde el ano 1; firmar entre 2030 y 2032')]
for lb, a, b in rows:
    last = lb == 'VEREDICTO'
    ws.cell(row=r, column=1, value=lb).font = F(10, True, GREEN if last else INK)
    ca = ws.cell(row=r, column=2, value=a)
    ca.font = F(9.5, last, RED if last else INK)
    cb = ws.cell(row=r, column=3, value=b)
    cb.font = F(9.5, last, GREEN if last else INK)
    for col in (2, 3):
        ws.cell(row=r, column=col).alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[r].height = 32
    r += 1
r += 1
para(ws, r, 'POR QUE EL PPA DE COMPRA NO MEJORA CON EL TIEMPO', 3, True, GREEN, 18)
r += 1
para(ws, r, 'Mientras persista la saturacion solar diurna, comprar la energia de carga a 28 USD/MWh seguira '
     'siendo peor que comprarla a spot. Lo que se reevalua en 2030 NO es esta oferta: es un instrumento '
     'distinto. Conviene decirlo asi al Comite, para no dejar la impresion de que el proyecto posterga una '
     'decision que en realidad ya tomo.', 3, h=44)
r += 2
para(ws, r, 'ADVERTENCIA SOBRE LA FASE 2 - hay que ajustar el lenguaje', 3, True, AMBER, 18)
r += 1
para(ws, r, 'La investigacion de mercado no encontro evidencia de PPA ni de tolling para almacenamiento a '
     'escala PMG (menos de 10 MW) en Chile. Los contratos verificados son de cientos de MW con contrapartes '
     'investment grade: Grenergy con una utility por 1 TWh/ano a 15 anos, y Grenergy con Codelco por 0,5 '
     'TWh/ano. Prometer al banco "un PPA en 2032" para 9 MW es una promesa que hoy no tiene mercado que la '
     'respalde.', 3, h=52)
r += 1
para(ws, r, 'Los instrumentos realistas a esta escala son otros y conviene nombrarlos asi: incorporacion a un '
     'portafolio de comercializacion o a un agregador; floor o collar nocturno con un comercializador; tolling '
     'con un actor mayor; o el regimen de precio estabilizado, este ultimo con la advertencia del DS 1/2026 '
     'que figura en la hoja Fuentes.', 3, h=44)

# ============================================================ 9. TRANSMISION
ws = wb.create_sheet('Transmision')
head(ws, 'La objecion de transmision, contestada con fuentes',
     'Objecion: "Kimal-Lo Aguirre traera energia del norte al sur y hundira el precio de venta nocturno". '
     'Cada respuesta esta anclada en fuente publica citable.', 4)
for col, w in zip('ABCD', [28, 52, 24, 40]):
    ws.column_dimensions[col].width = w
r = 4
tblhead(ws, r, ['Punto', 'Hecho verificado', 'Fecha / estado', 'Fuente'])
r += 1
tr = [('Punto de aterrizaje',
       'La linea va de S/E Kimal (Antofagasta) a S/E Lo Aguirre (Pudahuel, Region Metropolitana). NO llega a la Region del Maule. No existe tramo al sur en este proyecto.',
       'HVDC +-600 kV, 3.000 MW, 1.342 km', 'Coordinador Electrico Nacional; CNE 04-feb-2026'),
      ('Cronograma',
       'RCA aprobada 14-nov-2025. Construccion iniciada 4-feb-2026. Operacion comercial estimada mayo 2029; el Informe Final PET2026 de la CNE usa 2030 para el acoplamiento norte-centro.',
       '2029-2030, con historial de retrasos', 'CNE; ISA Interchile; La Tercera'),
      ('Que energia transporta',
       'Su objetivo declarado es evacuar excedente solar y eolico DIURNO del norte hacia el centro. La linea no almacena: transporta flujo en tiempo real.',
       'Verificado en la descripcion de la CNE', 'CNE; PET2026'),
      ('Efecto por franja horaria',
       'De noche el norte no tiene excedente solar que exportar. El efecto directo sobre el costo marginal NOCTURNO del centro-sur es estructuralmente menor que el diurno.',
       'INFERENCIA fisica. No hay cifra oficial', 'No cuantificado en fuente oficial'),
      ('Topologia del nodo',
       'S/E Panimavida es de 220 kV (linea Ancoa-San Ambrosio). El nodo del BESS, a 13,2 kV, esta aguas abajo. Del troncal de 500 kV lo separan dos escalones de transformacion.',
       'Verificado', 'CNE Calificacion de Instalaciones; Coordinador EAF 011/2019'),
      ('El HVDC que si cruza el Maule',
       'Lo Aguirre-Entre Rios (+-600 kV) pasa por el Maule, pero su proposito es evacuar excedente EOLICO DEL SUR hacia el centro-norte: flujo en sentido contrario al que preocupa. Fecha no verificada oficialmente.',
       'Adjudicacion condicionada; fecha NO VERIFICADA', 'Electromineria; Prieto Abogados; CNE ITP PET2024'),
      ('Desacople del sur',
       'La CNE proyecta que las diferencias de costo marginal en la zona sur PERSISTEN hacia 2032 incluso con las obras de 500 kV, sin acoplamiento pleno de los costos marginales anuales.',
       'Informe Final PET2026', 'CNE / Coordinador, PET2026'),
      ('Vertimiento creciente',
       'Vertimiento renovable ene-feb 2026: 1.402 GWh, un 29,3% mas que el ano anterior. Confirma que el excedente diurno sigue creciendo, es decir que el lado carga se sigue abaratando.',
       'ene-feb 2026', 'Renewables Now; PV-Tech / ACERA')]
for p, h_, f_, s in tr:
    ws.cell(row=r, column=1, value=p).font = F(10, True)
    for col, val in [(2, h_), (3, f_), (4, s)]:
        c = ws.cell(row=r, column=col, value=val)
        c.font = F(9 if col == 2 else 8.5, False, INK if col == 2 else GREY)
        c.alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[r].height = 48
    r += 1
r += 1
para(ws, r, 'EL GIRO DEL ARGUMENTO - esto es lo que hay que decirle al director', 4, True, GREEN, 18)
r += 1
for t in ['La preocupacion es que Kimal-Lo Aguirre baje el precio de venta nocturno. Pero los datos del propio '
          'nodo muestran que el precio de venta nocturno NO viene cayendo: esta plano en 98-100 USD/MWh desde hace '
          'tres anos. Lo que si se movio, y comprimio el spread, fue el COSTO DE CARGA diurno, que subio de 10,1 a '
          '22,1 USD/MWh.',
          'Y Kimal-Lo Aguirre actua precisamente sobre el costo marginal diurno, empujandolo A LA BAJA. Sobre la '
          'evidencia disponible, la linea que se teme es viento de cola para el riesgo real del proyecto, no viento '
          'de frente.',
          'El riesgo verdadero -el que identifico el asesor tecnico- es que el costo de carga suba a medida que baja '
          'el vertimiento y las baterias compiten por el excedente del mediodia. Ese riesgo esta explicitado en el '
          'gatillo 3.']:
    para(ws, r, t, 4, h=44)
    r += 1

# ============================================================ 10. FUENTES
ws = wb.create_sheet('Fuentes')
head(ws, 'Fuentes', 'Cada dato publicable de este libro tiene fuente. Lo marcado NO VERIFICADO no se publica.', 3)
for col, w in zip('ABC', [46, 28, 86]):
    ws.column_dimensions[col].width = w
r = 4
tblhead(ws, r, ['Dato', 'Fuente', 'URL'])
r += 1
fu = [('BESS en operacion en el SEN: 2.283 MW / 9.346 MWh (mar-2026)', 'Coordinador via Electromineria',
       'https://electromineria.cl/potencia-instalada-almacenamiento-sistema-electrico-se-duplicara-2026/'),
      ('BESS en construccion: 6.358 MW en 74 proyectos (jun-2026)', 'CNE via Actualidad Juridica',
       'https://actualidadjuridica.doe.cl/proyectos-de-almacenamiento-con-baterias-en-construccion-superan-los-6-300-mw-en-chile/'),
      ('Distribucion regional del BESS en construccion: Antofagasta 44%, Atacama 23%, Tarapaca 13,6%, RM 5%. El Maule NO figura',
       'CNE (jun-2026)', 'mismo enlace anterior'),
      ('NEM Australia: spread -85% en un ano (AU$342 a AU$51/MWh) con flota BESS de 4.360 a 9.000 MW',
       'Energy-Storage.News 29-jul-2026',
       'https://www.energy-storage.news/nem-battery-price-spreads-fall-85-in-a-year-as-australias-grid-scale-bess-fleet-passes-9000mw/'),
      ('CAISO: BESS sobre 13 GW y spread TB4 proyectado AL ALZA, a US$240-270/MWh hacia 2030', 'Modo Energy feb-2026',
       'https://modoenergy.com/research/en/caiso-market-outlook-february-2026-california-power-price-spread'),
      ('Kimal-Lo Aguirre: +-600 kV, 3.000 MW, de Kimal a Lo Aguirre (RM)', 'Coordinador Electrico Nacional',
       'https://www.coordinador.cl/desarrollo/documentos/grandes-proyectos/proyecto-hvdc-kimal-lo-aguirre/'),
      ('Inicio de construccion del HVDC: 4-feb-2026', 'CNE',
       'https://www.cne.cl/prensa/prensa-2026/02-febrero-2026/cne-participa-en-inicio-de-la-etapa-de-construccion-de-la-linea-de-transmision-hvdc-kimal-lo-aguirre/'),
      ('El desacople del sur persiste hacia 2032 incluso con las obras de 500 kV', 'CNE, Informe Final PET2026',
       'https://www.coordinador.cl/wp-content/uploads/2026/01/Informe-Final-PET2026.pdf'),
      ('Vertimiento ene-feb 2026: 1.402 GWh, +29,3% interanual', 'Renewables Now',
       'https://renewablesnow.com/news/chiles-renewables-curtailment-already-at-1-4-twh-in-jan-feb-2026-1292010/'),
      ('NTCO-PMGD del 19-feb-2026, Capitulo 9: coordinacion de PMGD CON almacenamiento con el Coordinador', 'CNE',
       'https://www.cne.cl/wp-content/uploads/2026/02/2026.02.19_NTCO-PMGD-2026.pdf'),
      ('ATENCION - DS 1/2026: incorporar BESS a un PMGD con precio estabilizado se considera "modificacion" y puede hacer PERDER el precio estabilizado',
       'Cuatrecasas',
       'https://www.cuatrecasas.com/es/latam/art/energia-infraestructura-pmgd-almacenamiento-alcances-reingreso-ds-1-2026'),
      ('PPA nocturno de BESS por 1 TWh/ano a 15 anos (escala de cientos de MW, no PMG)', 'PV Magazine LatAm 30-jun-2026',
       'https://www.pv-magazine-latam.com/2026/06/30/en-chile-grenergy-firma-su-mayor-ppa-para-suministro-nocturno-desde-baterias/'),
      ('Datos horarios del nodo BA S/E Panimavida 13,2 kV (BP1): 38.064 horas, del 2022-01-01 al 2026-05-22',
       'Coordinador Electrico Nacional, Real Definitivo', 'Descarga directa del Coordinador'),
      ('NO VERIFICADO - no publicar: "pipeline total sobre 36.000 MW"; "solo 7-10% del pipeline operativo"; "85% de las obras de transmision retrasadas"',
       '-', 'Solo en fuentes secundarias e inconsistente con las cifras oficiales')]
for d_, s, u in fu:
    warn = d_.startswith('ATENCION') or d_.startswith('NO VERIFICADO')
    c = ws.cell(row=r, column=1, value=d_)
    c.font = F(9, warn, AMBER if warn else INK)
    c.alignment = Alignment(wrap_text=True, vertical='top')
    c2 = ws.cell(row=r, column=2, value=s)
    c2.font = F(8.5, c=GREY)
    c2.alignment = Alignment(wrap_text=True, vertical='top')
    c3 = ws.cell(row=r, column=3, value=u)
    c3.font = F(8, c=BLUE)
    c3.alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[r].height = 32
    r += 1

# ====================== PATCH FINAL: conectar el Panel a sus fuentes reales ==================
# Se hace al final porque las direcciones destino solo existen una vez construidas las hojas.
assert SRC_BE == BE_ADDR, f'La direccion del break-even cambio: {SRC_BE} != {BE_ADDR}'
wsp = wb['Panel']
for row, src, fmt in [(PANEL_CARD_ROWS[0], SRC_MARGEN, '#,##0'),
                      (PANEL_CARD_ROWS[1], SRC_BE, '0.0'),
                      (PANEL_CARD_ROWS[2], SRC_CRUCE, '0')]:
    c = wsp.cell(row=row, column=2, value=f'={src}')
    c.font = F(15, True, GREEN)
    c.number_format = fmt
    c.alignment = Alignment(horizontal='center')
print('Panel conectado a:', SRC_MARGEN, '|', SRC_BE, '|', SRC_CRUCE)

out = (r'C:\Users\nicol\OneDrive\Documentos\0.2.Rho\Banco_Panimavida'
       r'\entrega_banco_v8\Panimavida_Ventana_Arbitraje_v8.xlsx')
os.makedirs(os.path.dirname(out), exist_ok=True)
wb.save(out)
print('OK ->', out)
print('hojas:', wb.sheetnames)
