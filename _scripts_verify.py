# -*- coding: utf-8 -*-
"""Recalcula el libro con LibreOffice headless y verifica los valores resultantes."""
import os, sys, shutil, subprocess, glob
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

SRC = (r'C:\Users\nicol\OneDrive\Documentos\0.2.Rho\Banco_Panimavida'
       r'\entrega_banco_v8\Panimavida_Ventana_Arbitraje_v8.xlsx')
WORK = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(WORK, 'lo_out')
SOF = r'C:\Program Files\LibreOffice\program\soffice.exe'

shutil.rmtree(OUT, ignore_errors=True)
os.makedirs(OUT, exist_ok=True)
tmp = os.path.join(WORK, 'verify_in.xlsx')
shutil.copy(SRC, tmp)

env = dict(os.environ)
p = subprocess.run([SOF, '--headless', '--norestore', '--invisible',
                    '--convert-to', 'xlsx:Calc MS Excel 2007 XML',
                    '--outdir', OUT, tmp],
                   capture_output=True, text=True, timeout=300, env=env)
print('soffice rc=', p.returncode)
if p.stdout.strip():
    print(p.stdout.strip()[:400])
if p.stderr.strip():
    print('stderr:', p.stderr.strip()[:400])

files = glob.glob(os.path.join(OUT, '*.xlsx'))
if not files:
    print('!! LibreOffice no produjo salida')
    sys.exit(1)
print('recalculado ->', os.path.basename(files[0]))

wb = openpyxl.load_workbook(files[0], data_only=True)

def show(sheet, label, cells):
    ws = wb[sheet]
    print(f'\n--- {label} ({sheet}) ---')
    for addr in cells:
        v = ws[addr].value
        print(f'  {addr:>6} = {v!r}')

# Supuestos: formulas derivadas
ws = wb['Supuestos']
print('\n=== SUPUESTOS: formulas derivadas ===')
for row in range(4, 30):
    lbl = ws.cell(row=row, column=1).value
    if lbl in ('Energia de venta por ciclo', 'Spread', 'Costo de carga (nivel)',
               'Precio de venta 4 h punta (nivel)', 'Energia de carga por ciclo',
               'Eficiencia round-trip'):
        print(f'  {lbl:<42} = {ws.cell(row=row, column=2).value!r}')

# Escenarios: proyeccion
ws = wb['Escenarios']
print('\n=== ESCENARIOS: primeras filas de cada bloque ===')
for row in range(1, ws.max_row + 1):
    a = ws.cell(row=row, column=1).value
    if isinstance(a, str) and a.startswith('Escenario ') and len(a) < 14:
        vals = [ws.cell(row=row, column=j).value for j in range(2, 11)]
        vals = [round(v, 1) if isinstance(v, (int, float)) else v for v in vals]
        print(f'  fila {row:>3} {a}: {vals}')

# Margen
ws = wb['Margen_DSCR']
print('\n=== MARGEN_DSCR ===')
for row in range(1, ws.max_row + 1):
    a = ws.cell(row=row, column=1).value
    if isinstance(a, str) and a.strip() and not a.startswith('B.') and not a.startswith('A.'):
        b = ws.cell(row=row, column=2).value
        if b is not None and not isinstance(b, str):
            print(f'  {a[:52]:<52} = {b if not isinstance(b,float) else round(b,2)!r}')
        elif isinstance(b, str) and b:
            print(f'  {a[:52]:<52} = {b!r}')
print('  filas de margen proyectado:')
for row in range(1, ws.max_row + 1):
    a = ws.cell(row=row, column=1).value
    if isinstance(a, str) and a.startswith('Escenario ') and len(a) < 14:
        vals = [ws.cell(row=row, column=j).value for j in range(2, 11)]
        vals = [int(v) if isinstance(v, (int, float)) else v for v in vals]
        print(f'    {a}: {vals}')
print('  celdas puente al Panel:')
print('    C7  =', ws['C7'].value)
print('    C30 =', ws['C30'].value)

# Break even
ws = wb['Break_even']
print('\n=== BREAK_EVEN ===')
for row in range(5, 14):
    a = ws.cell(row=row, column=1).value
    if isinstance(a, (int, float)):
        vals = [ws.cell(row=row, column=j).value for j in range(2, 7)]
        vals = [round(v, 1) if isinstance(v, (int, float)) else v for v in vals]
        print(f'  servicio {int(a):>7,}: 182c..365c = {vals}')
print('  C8 (referencia del Panel) =', ws['C8'].value)

# Panel
ws = wb['Panel']
print('\n=== PANEL: las tres cifras ===')
for row in range(1, 30):
    a = ws.cell(row=row, column=1).value
    if isinstance(a, str) and ('Margen bruto' in a or 'Spread necesario' in a or 'Ano en que' in a):
        print(f'  {a[:50]:<50} = {ws.cell(row=row, column=2).value!r}')
print('\nOK')
